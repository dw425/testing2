"""
Excel report generation for LHO Lite.

Returns BytesIO objects for Flask send_file() — no temp files needed.
Universal across AWS / Azure / GCP.
"""

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')
ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')


def _clean(v):
    if isinstance(v, str):
        v = ANSI_RE.sub('', v)
        v = ILLEGAL_RE.sub('', v)
    return v


def _fmt_bytes(b):
    if not b:
        return "—"
    b = int(b)
    if b < 1024:
        return f"{b} B"
    if b < 1048576:
        return f"{b / 1024:.1f} KB"
    if b < 1073741824:
        return f"{b / 1048576:.2f} MB"
    return f"{b / 1073741824:.2f} GB"


# Shared styles
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_CRIT_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
_CRIT_FONT = Font(bold=True, color="FFFFFF")
_HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_LOW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_BDR = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")
_SEV_FILLS = {"CRITICAL": _CRIT_FILL, "HIGH": _HIGH_FILL, "MEDIUM": _MED_FILL, "LOW": _LOW_FILL}


def _safe_int(v):
    try: return int(v)
    except (TypeError, ValueError): return 0


def _safe_float(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0


def _style_hdr(ws, row, cols, color="1F4E79"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HDR_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BDR


def _auto_width(ws, max_w=55):
    for col in ws.columns:
        m = 0
        cl = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                m = max(m, min(len(str(cell.value).split("\n")[0]), max_w))
        ws.column_dimensions[cl].width = max(m + 2, 14)


# ---------------------------------------------------------------------------
# Security / FedRAMP Excel
# ---------------------------------------------------------------------------

def generate_security_excel(sec_data: dict, findings: list) -> io.BytesIO:
    """Generate Security & FedRAMP Excel report.  Returns BytesIO."""
    wb = Workbook()
    cloud = sec_data.get("_cloud", "Unknown")

    # Sheet 1: Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = f"Databricks Workspace Security & FedRAMP Analysis ({cloud})"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(italic=True, size=10, color="808080")

    me = sec_data.get("me", {})
    users = sec_data.get("users", {})
    group_list = sec_data.get("groups", {}).get("Resources", [])
    admin_group = next((g for g in group_list if g.get("displayName") == "admins"), {})
    cluster_list = sec_data.get("clusters", {}).get("clusters", [])
    warehouse_list = sec_data.get("warehouses", {}).get("warehouses", [])
    catalog_list = sec_data.get("catalogs", {}).get("catalogs", [])
    job_list = sec_data.get("jobs", {}).get("jobs", [])

    r = 4
    for c, h in enumerate(["Property", "Value"], 1):
        ws.cell(row=r, column=c, value=h)
    _style_hdr(ws, r, 2)

    overview = [
        ("Cloud", cloud),
        ("Total Users", str(users.get("totalResults", 0))),
        ("Admin Users", str(len(admin_group.get("members", [])))),
        ("Clusters", str(len(cluster_list))),
        ("SQL Warehouses", str(len(warehouse_list))),
        ("Catalogs", ", ".join(c.get("name", "") for c in catalog_list)),
        ("Jobs", str(len(job_list))),
        ("Total Findings", str(len(findings))),
        ("Critical", str(sum(1 for f in findings if f[0] == "CRITICAL"))),
        ("High", str(sum(1 for f in findings if f[0] == "HIGH"))),
    ]
    for prop, val in overview:
        r += 1
        ws.cell(row=r, column=1, value=prop)
        ws.cell(row=r, column=2, value=_clean(val))
    _auto_width(ws)

    # Sheet 2: Findings
    ws2 = wb.create_sheet("Security Findings")
    h2 = ["ID", "Severity", "Category", "NIST", "Finding", "Impact", "Recommendation"]
    for c, h in enumerate(h2, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_hdr(ws2, 1, len(h2))
    for i, (sev, cat, nist, finding, impact, rec) in enumerate(findings, 1):
        r2 = i + 1
        ws2.cell(row=r2, column=1, value=f"SEC-{i:03d}")
        ws2.cell(row=r2, column=2, value=sev)
        ws2.cell(row=r2, column=3, value=cat)
        ws2.cell(row=r2, column=4, value=nist)
        ws2.cell(row=r2, column=5, value=_clean(finding))
        ws2.cell(row=r2, column=6, value=_clean(impact))
        ws2.cell(row=r2, column=7, value=_clean(rec))
        if sev in _SEV_FILLS:
            ws2.cell(row=r2, column=2).fill = _SEV_FILLS[sev]
        if sev == "CRITICAL":
            ws2.cell(row=r2, column=2).font = _CRIT_FONT
    _auto_width(ws2)

    # Sheet 3: FedRAMP Roadmap
    ws3 = wb.create_sheet("FedRAMP Roadmap")
    h3 = ["Priority", "Phase", "Action", "NIST", "Effort", "Status"]
    for c, h in enumerate(h3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_hdr(ws3, 1, len(h3))
    r3 = 1
    sev_index = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    for sev, cat, nist, finding, impact, rec in findings:
        r3 += 1
        phase = {"CRITICAL": "Week 1", "HIGH": "Week 1-2", "MEDIUM": "Week 2-4"}.get(sev, "Month 2-3")
        ws3.cell(row=r3, column=1, value=f"P{sev_index.get(sev, 3)}")
        ws3.cell(row=r3, column=2, value=phase)
        ws3.cell(row=r3, column=3, value=_clean(rec))
        ws3.cell(row=r3, column=4, value=nist)
        ws3.cell(row=r3, column=5, value="Hours-days")
        ws3.cell(row=r3, column=6, value="NOT STARTED")
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Usage & Cost Excel
# ---------------------------------------------------------------------------

def generate_usage_excel(sec_data: dict, usage_data: dict) -> io.BytesIO:
    """Generate Usage & Cost Excel report.  Returns BytesIO."""
    wb = Workbook()
    cloud = sec_data.get("_cloud", "Unknown")

    apps = sec_data.get("apps", {}).get("apps", [])
    models = sec_data.get("serving", {}).get("endpoints", [])
    daily_rows = usage_data.get("daily_queries", {}).get("rows", [])
    user_rows = usage_data.get("user_queries", {}).get("rows", [])
    wh_rows = usage_data.get("warehouse_events", {}).get("rows", [])
    price_rows = usage_data.get("list_prices", {}).get("rows", [])
    schema_rows = usage_data.get("schema_overview", {}).get("rows", [])

    # Sheet 1: Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = f"Databricks Usage & Cost Report ({cloud})"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"

    total_queries = sum(_safe_int(r[1]) for r in daily_rows) if daily_rows else 0
    total_succeeded = sum(_safe_int(r[2]) for r in daily_rows) if daily_rows else 0
    total_gb = round(sum(_safe_float(r[4]) for r in daily_rows), 2) if daily_rows else 0
    total_min = round(sum(_safe_float(r[6]) for r in daily_rows), 1) if daily_rows else 0
    total_tables = sum(_safe_int(r[2]) for r in schema_rows if r[1] != "information_schema") if schema_rows else 0

    summary = [
        ("Cloud", cloud),
        ("Active Apps", str(len(apps))),
        ("Model Endpoints", str(len(models))),
        ("Total User Tables", str(total_tables)),
        ("Total Queries (30d)", f"{total_queries:,}"),
        ("Success Rate", f"{total_succeeded * 100 // max(total_queries, 1)}%"),
        ("Data Read (30d)", f"{total_gb} GB"),
        ("Compute Time (30d)", f"{total_min} min"),
    ]
    for i, (k, v) in enumerate(summary, 4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    _auto_width(ws)

    # Sheet 2: Apps & Models
    ws2 = wb.create_sheet("Apps & Models")
    headers = ["Name", "Description", "State", "Compute", "Creator", "Created"]
    for j, h in enumerate(headers, 1):
        ws2.cell(row=1, column=j, value=h)
    _style_hdr(ws2, 1, len(headers))
    for i, app in enumerate(apps, 2):
        ws2.cell(row=i, column=1, value=app.get("name", ""))
        ws2.cell(row=i, column=2, value=_clean(app.get("description", "")))
        state = app.get("compute_status", {})
        ws2.cell(row=i, column=3, value=state.get("state", "") if isinstance(state, dict) else "")
        ws2.cell(row=i, column=4, value=app.get("compute_size", ""))
        ws2.cell(row=i, column=5, value=app.get("creator", ""))
        ws2.cell(row=i, column=6, value=(app.get("create_time", "") or "")[:10])

    r = len(apps) + 4
    mh = ["Model", "Input DBU/1M", "Output DBU/1M", "Input $/1M", "Output $/1M"]
    for j, h in enumerate(mh, 1):
        ws2.cell(row=r, column=j, value=h)
    _style_hdr(ws2, r, len(mh), "2E75B6")
    for i, ep in enumerate(models, r + 1):
        entities = ep.get("config", {}).get("served_entities", [{}])
        fm = entities[0].get("foundation_model", {}) if entities else {}
        inp = float(fm.get("input_price", "0") or "0")
        out = float(fm.get("price", "0") or "0")
        ws2.cell(row=i, column=1, value=fm.get("display_name", ep.get("name", "")))
        ws2.cell(row=i, column=2, value=inp)
        ws2.cell(row=i, column=3, value=out)
        ws2.cell(row=i, column=4, value=round(inp * 0.07, 4))
        ws2.cell(row=i, column=5, value=round(out * 0.07, 4))
    _auto_width(ws2)

    # Sheet 3: Table Inventory
    ws3 = wb.create_sheet("Table Inventory")
    th = ["Table", "Schema", "Type", "Created", "Description", "Size"]
    for j, h in enumerate(th, 1):
        ws3.cell(row=1, column=j, value=h)
    _style_hdr(ws3, 1, len(th), "548235")
    r3 = 1
    for key, inv in usage_data.get("table_inventory", {}).items():
        parts = key.split(".")
        schema = parts[-1] if len(parts) > 1 else key
        for row in inv.get("rows", []):
            r3 += 1
            ws3.cell(row=r3, column=1, value=row[0])
            ws3.cell(row=r3, column=2, value=schema)
            ws3.cell(row=r3, column=3, value=row[1])
            ws3.cell(row=r3, column=4, value=(row[2] or "")[:10] if row[2] else "")
            ws3.cell(row=r3, column=5, value=_clean((row[4] or "")[:80]) if len(row) > 4 else "")
            fqn = f"{key}.{row[0]}"
            size = usage_data.get("table_sizes", {}).get(fqn)
            ws3.cell(row=r3, column=6, value=_fmt_bytes(size) if size else "")
    _auto_width(ws3)

    # Sheet 4: Query by User
    ws4 = wb.create_sheet("Query Usage by User")
    uh = ["User", "Status", "Queries", "Data Read (GB)", "Rows", "Duration (min)"]
    for j, h in enumerate(uh, 1):
        ws4.cell(row=1, column=j, value=h)
    _style_hdr(ws4, 1, len(uh), "BF8F00")
    for i, r in enumerate(user_rows, 2):
        ws4.cell(row=i, column=1, value=r[0])
        ws4.cell(row=i, column=2, value=r[1])
        ws4.cell(row=i, column=3, value=_safe_int(r[2]))
        ws4.cell(row=i, column=4, value=_safe_float(r[3]))
        ws4.cell(row=i, column=5, value=_safe_int(r[4]))
        ws4.cell(row=i, column=6, value=_safe_float(r[5]))
    _auto_width(ws4)

    # Sheet 5: Daily Trends
    ws5 = wb.create_sheet("Daily Trends")
    dh = ["Date", "Total", "Succeeded", "Failed", "Read (GB)", "Rows", "Compute (min)"]
    for j, h in enumerate(dh, 1):
        ws5.cell(row=1, column=j, value=h)
    _style_hdr(ws5, 1, len(dh), "7030A0")
    for i, r in enumerate(daily_rows, 2):
        for j, v in enumerate(r, 1):
            ws5.cell(row=i, column=j, value=_safe_float(v) if j > 1 else v)
    _auto_width(ws5)

    # Sheet 6: Warehouse Events
    ws6 = wb.create_sheet("Warehouse Events")
    weh = ["Warehouse ID", "Event", "Time", "Clusters"]
    for j, h in enumerate(weh, 1):
        ws6.cell(row=1, column=j, value=h)
    _style_hdr(ws6, 1, len(weh), "C00000")
    for i, r in enumerate(wh_rows, 2):
        ws6.cell(row=i, column=1, value=r[0])
        ws6.cell(row=i, column=2, value=r[1])
        ws6.cell(row=i, column=3, value=r[2])
        ws6.cell(row=i, column=4, value=_safe_int(r[3]))
    _auto_width(ws6)

    # Sheet 7: DBU Pricing
    ws7 = wb.create_sheet("DBU Pricing")
    ph = ["SKU", "Price (USD)", "Unit"]
    for j, h in enumerate(ph, 1):
        ws7.cell(row=1, column=j, value=h)
    _style_hdr(ws7, 1, len(ph), "808080")
    rp = 1
    for r in price_rows:
        rp += 1
        ws7.cell(row=rp, column=1, value=r[0])
        ws7.cell(row=rp, column=2, value=_safe_float(r[1]))
        ws7.cell(row=rp, column=3, value=r[2])
    _auto_width(ws7)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
